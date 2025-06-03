import io
from datetime import date
import openpyxl
from openpyxl.utils import get_column_letter

def create_monthly_report_excel(
    username: str,
    anno: int,
    mese: int,
    daily_details: list,
    summary: dict
) -> io.BytesIO:
    """
    daily_details: lista di dict con chiavi: date (date), day_type (str), punches_formatted (str),
                   ore_ordinarie (float), ore_extra (float), note (str)
    summary: dict con totali: total_worked_days, total_ore_ordinarie, total_ore_extra, ferie, permessi, malattia.
    Ritorna BytesIO con file xlsx.
    """
    wb = openpyxl.Workbook()
    ws_det = wb.active
    ws_det.title = "Dettaglio Giorni"

    headers = ["Data", "Tipo Giornata", "Ingressi/Uscite", "Ore Ordinarie", "Ore Extra", "Note"]
    ws_det.append(headers)
    for row in daily_details:
        ws_det.append([
            row["date"].strftime("%Y-%m-%d"),
            row["day_type"],
            row["punches_formatted"],
            row["ore_ordinarie"],
            row["ore_extra"],
            row.get("note", "")
        ])
    # Regolazioni colonne
    for idx, _ in enumerate(headers, 1):
        ws_det.column_dimensions[get_column_letter(idx)].width = 18

    # Foglio riepilogo
    ws_sum = wb.create_sheet("Riepilogo Mensile")
    ws_sum.append(["Utente", username])
    ws_sum.append(["Mese", f"{anno}-{mese:02d}"])
    ws_sum.append([])
    ws_sum.append(["Totale Giorni Lavorati", summary["total_worked_days"]])
    ws_sum.append(["Totale Ore Ordinarie", summary["total_ore_ordinarie"]])
    ws_sum.append(["Totale Ore Straordinarie", summary["total_ore_extra"]])
    ws_sum.append(["Giorni Ferie", summary["ferie"]])
    ws_sum.append(["Giorni Permessi", summary["permessi"]])
    ws_sum.append(["Giorni Malattia", summary["malattia"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
